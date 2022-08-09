import sys, os, datetime


RABBIT_QUEUE = os.getenv('RABBIT_QUEUE', '')
RABBIT_HOST = os.getenv('RABBIT_HOST', 'spark-rabbitmq')
RABBIT_USER = os.getenv('RABBIT_USER', '')
RABBIT_PASSWORD = os.getenv('RABBIT_PASSWORD', '')

SQL_HOST = os.getenv('SQL_HOST', '')
SQL_USER = os.getenv('SQL_USER', '')
SQL_DRIVER = os.getenv('SQL_DRIVER', '')
SQL_DRIVER_SRC = os.getenv('SQL_DRIVER_SRC', '')

FOLDER_RAW = os.getenv('WATCH_FOLDER_DATA', '/temp/raw/')
FOLDER_TEMP = os.getenv('WATCH_FOLDER_TEMP', '/temp/raw/')
FOLDER_INPUT = os.getenv('WATCH_FOLDER_INPUT', '/data/input/')

def get_fullname(directory, filename, extension):
    return "{}{}.{}".format(directory, str(filename), extension)

def normalize_xls(filename):
    import jpype, uuid, shutil, asposecells    
    from subprocess import  Popen
 
    jpype.startJVM()
    from asposecells.api import Workbook
  
    temp_filename = "{}".format(str(uuid.uuid4())) 
    
    workbook = Workbook(get_fullname(FOLDER_RAW, filename,"xls"))    
    workbook.save(get_fullname(FOLDER_TEMP, temp_filename,"xlsx"))

    jpype.shutdownJVM()

    p = Popen(['libreoffice', '--headless', '--convert-to', 'xls', '--outdir',
               FOLDER_TEMP, get_fullname(FOLDER_TEMP, temp_filename,"xlsx")]) 
    p.communicate()

    p = Popen(['libreoffice', '--headless', '--convert-to', 'xlsx', '--outdir',
               FOLDER_TEMP, get_fullname(FOLDER_TEMP, temp_filename,"xls")]) 
    p.communicate()
   
    shutil.move(get_fullname(FOLDER_TEMP, temp_filename,"xlsx"),get_fullname(FOLDER_INPUT, filename,"xlsx"))

def get_workbook(filename):  
    from openpyxl import load_workbook

    workbook = load_workbook(get_fullname(FOLDER_INPUT, filename,"xlsx"))
    return workbook

def get_pivot_cache(worksheet, pivot_name):
    import numpy as np  
    from openpyxl.pivot.fields import Missing

    fields_map = {}
    rows = []

    pivot_table = [p for p in worksheet._pivots if p.name == pivot_name][0]
    
    for field in pivot_table.cache.cacheFields:
        if field.sharedItems.count > 0:
            fields_map[field.name] = [field.v if not isinstance(field, Missing) else np.nan for field in field.sharedItems._fields]

    column_names = [field.name for field in pivot_table.cache.cacheFields]
    
    record = pivot_table.cache.records.r[0]   
    record_values = [
        field.v if not isinstance(field, Missing) else np.nan for field in record._fields
    ]

    for record in pivot_table.cache.records.r:   
        record_values = [
            field.v if not isinstance(field, Missing) else np.nan for field in record._fields
        ]

        row_dict = {k: v for k, v in zip(column_names, record_values)}
     
        for key in fields_map:
            row_dict[key] = fields_map[key][row_dict[key]]

        rows.append(row_dict)
    
    return rows

def convert_rdd_table(spark, tableRDD):
    from pyspark.sql.functions import lit, unix_timestamp, expr , monotonically_increasing_id   
    from pyspark.sql.types import StructType, StructField, StringType, IntegerType, DoubleType, TimestampType 
    import time
    import datetime

    created_at = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')
    unpivotExpr = "stack(12, '01-01', Jan ,'02-01', Fev,'03-01', Mar,'04-01', Abr,'05-01', Mai,'06-01', Jun, '07-01', Jul,'08-01', Ago,'09-01', Set,'10-01', Out,'11-01', Nov,'12-01', Dez) as (month,volume)"

    schema = StructType([
        StructField('COMBUSTÍVEL', StringType(), True),
        StructField('ANO', DoubleType(), True), 
        StructField('ESTADO', StringType(), True),
        StructField('UNIDADE', StringType(), True),
        StructField('Jan', DoubleType(), True),
        StructField('Fev', DoubleType(), True),
        StructField('Mar', DoubleType(), True),
        StructField('Abr', DoubleType(), True),
        StructField('Mai', DoubleType(), True),
        StructField('Jun', DoubleType(), True),
        StructField('Jul', DoubleType(), True),
        StructField('Ago', DoubleType(), True),
        StructField('Set', DoubleType(), True),
        StructField('Out', DoubleType(), True),
        StructField('Nov', DoubleType(), True),
        StructField('Dez', DoubleType(), True),
        StructField('TOTAL', DoubleType(), True),
        StructField('created_at', TimestampType(), True)
    ])

    tableDf = spark.createDataFrame(tableRDD, schema) \
                           .withColumnRenamed("COMBUSTÍVEL","product") \
                           .withColumnRenamed("ANO","year") \
                           .withColumnRenamed("UNIDADE","unit") \
                           .withColumnRenamed("ESTADO","uf") \
                           .withColumnRenamed("TOTAL","total")

    tableDf = tableDf.fillna(0, subset=['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'])
    tableDf.createOrReplaceTempView("Temporary") 

    tableRaw = spark.sql("select product "
                    + "        , year"
                    + "        , unit" 
                    + "        , uf" 
                    + "        , case when Jan+Fev+Mar+Abr+Mai+Jun+Jul+Ago+Set+Out+Nov+Dez == total then 1 else 0 end validy" 
                    + "        , Jan,Fev,Mar,Abr,Mai,Jun,Jul,Ago,Set,Out,Nov,Dez,total"
                    + " from Temporary ")
    
    tableFinalErrors = tableRaw.filter("validy == 0")
    
    tableRef = tableRaw.filter("validy == 1")

    tableRef.select("product","unit", "uf", "year", "validy", expr(unpivotExpr)) \
           .withColumn("id", monotonically_increasing_id()) \
           .createOrReplaceTempView("Temporary") 

    tableRefFinal = spark.sql("select  id"
                    + "              , cast(cast(year as int) ||'-'|| month as date) year_month"
                    + "              , uf" 
                    + "              , product" 
                    + "              , unit" 
                    + "              , volume "
                    + "     from Temporary " 
                    + "     where validy == 1 ") \
                    .withColumn('created_at',unix_timestamp(lit(created_at),'yyyy-MM-dd HH:mm:ss').cast("timestamp"))   

    return tableRefFinal , tableFinalErrors

def process_message(filename):    
    from os.path import exists    
    from pyspark.sql import SparkSession

    url = SQL_HOST
    properties = {
        "user": SQL_USER,
        "driver": SQL_DRIVER
    }

    spark = SparkSession \
            .builder \
            .config("spark.jars", SQL_DRIVER_SRC) \
            .appName("Spark Extract") \
            .getOrCreate() 

    worksheet_name = "Plan1"

    file_fullname = get_fullname(FOLDER_RAW, filename,"xls")
    
    print("arquivo recebido: {}".format(file_fullname))
    
    file_exists = exists(file_fullname)       
    if not file_exists: sys.exit() 

    normalize_xls(filename) 
 
    workbook = get_workbook(filename)
    worksheet = workbook['Plan1']
    
    vendaDosCombustiveisDerivadosDePetrolioCache = get_pivot_cache(worksheet, 'Tabela dinâmica1')
    vendaDeOleoDieselCache = get_pivot_cache(worksheet, 'Tabela dinâmica3')
    
    vendaDosCombustiveisDerivadosDePetrolioDF,vendaDosCombustiveisDerivadosDePetrolioDFErros = convert_rdd_table(spark, vendaDosCombustiveisDerivadosDePetrolioCache)
    vendaDeOleoDieselDF, vendaDeOleoDieselDFErrors  = convert_rdd_table(spark, vendaDeOleoDieselCache)

    vendaDosCombustiveisDerivadosDePetrolioDF.printSchema()
    vendaDosCombustiveisDerivadosDePetrolioDF.show()

    vendaDeOleoDieselDF.printSchema()
    vendaDeOleoDieselDF.show()
   
    vendaDosCombustiveisDerivadosDePetrolioDF.write \
    .jdbc(url=url, table="report_combustiveis_derivados", mode='append', properties=properties)

    vendaDeOleoDieselDF.write \
    .jdbc(url=url, table="report_oleo_diesel", mode='append', properties=properties)


    vendaDosCombustiveisDerivadosDePetrolioDFErros.write \
    .jdbc(url=url, table="report_combustiveis_derivados_errors", mode='append', properties=properties)

    vendaDeOleoDieselDFErrors.write \
    .jdbc(url=url, table="report_oleo_diesel_errors", mode='append', properties=properties)

def on_receive():
    import message as builder
    server = {
        'host': RABBIT_HOST,
        'user': RABBIT_USER,
        'pass': RABBIT_PASSWORD,
        'queue':RABBIT_QUEUE
    }

    def on_message(data):
        process_message(data)

    builder.enqueue(server).receive(on_message)

if __name__ == "__main__": 
     on_receive()